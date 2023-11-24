import zipfile
import pandas as pd
import openpyxl
from datetime import datetime
from Envia_Anexo import Envia_Lista

date=datetime.today().strftime('%Y%m%d')

# Nome do arquivo compactado
zip_r1 = "X:\\Help Desk\\Projeto Oi BC\\Bases Servtel\\R1\\SERVTEL_R1_"+str(date)+".zip"
zip_r2 = "X:\\Help Desk\\Projeto Oi BC\\Bases Servtel\\R2\\SERVTEL_R2_"+str(date)+".zip"

# Abrir o arquivo ZIP em modo de leitura
with zipfile.ZipFile(zip_r1, 'r') as zip_ref:
    # Listar os arquivos contidos no arquivo ZIP
    print(zip_ref.namelist())

    # Extrair um arquivo específico do arquivo ZIP
    zip_ref.extract("SERVTEL_R1_"+str(date)+".csv", "./")  # Segundo argumento é o diretório de extração
    print("Arquivo extraído com sucesso.")

with zipfile.ZipFile(zip_r2, 'r') as zip_ref:
    # Listar os arquivos contidos no arquivo ZIP
    print(zip_ref.namelist())

    # Extrair um arquivo específico do arquivo ZIP
    zip_ref.extract("SERVTEL_R2_"+str(date)+".csv", "./")  # Segundo argumento é o diretório de extração
    print("Arquivo extraído com sucesso.")

r1 = pd.read_csv("SERVTEL_R1_"+str(date)+".csv", sep=";",low_memory=False)
r2 = pd.read_csv("SERVTEL_R2_"+str(date)+".csv", sep=";",low_memory=False)
r_con = pd.concat([r1,r2])

book = openpyxl.load_workbook('Evolução Indicadores.xlsx')
sheet = book.active

date2=datetime.today().strftime('%Y-%m-%d 00:00:00')
terminal=sheet.cell(row=9, column=3).value
data_coluna=sheet.cell(row=7, column=1).value
try:
    data_coluna=data_coluna.strftime('%Y-%m-%d 00:00:00')
except:
    pass
i=0
j=0


while(data_coluna!=date2):
    print(data_coluna)
    data_coluna=sheet.cell(row=7, column=1+j).value
    try:
        data_coluna=data_coluna.strftime('%Y-%m-%d 00:00:00')
    except:
        pass
    j=j+1
    

while(terminal!=None):
    print(terminal)
    try:
        status=r_con[r_con['Terminal'] == terminal].STATUS_TUP.values[0]
        print(status)
        if status == 'ATIVO':
            sheet.cell(row=9+i, column=j).value = 1
        else:
            sheet.cell(row=9+i, column=j).value = status
        
    except:
        print('0')
        sheet.cell(row=9+i, column=j).value = 0
    i=i+1    
    terminal=sheet.cell(row=9+i, column=3).value
    

book.save('Evolução Indicadores.xlsx')
Envia_Lista()
